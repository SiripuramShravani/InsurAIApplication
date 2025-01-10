import random
import os
from django.conf import settings
from django.template.loader import render_to_string
from typing import ClassVar, Dict, Type, List, Optional, Union, Any
import base64
from django.core.mail import send_mail
from django.utils.html import strip_tags
from datetime import datetime, timedelta, timezone
from .master_package_databases import MongoDB
from .master_package_security import Authentication
from bson.objectid import ObjectId
import uuid
import json
import usaddress
import pandas as pd
import docx
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
import requests
import fitz
import googlemaps
import boto3
from botocore.exceptions import ClientError
from dotenv import load_dotenv
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_core.output_parsers import JsonOutputParser
from langchain_groq import ChatGroq
from langchain_core.prompts import PromptTemplate
from io import BytesIO
from langchain.output_parsers import PydanticOutputParser
load_dotenv()

AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
AWS_REGION = os.getenv('AWS_REGION')
groq_api_key = os.getenv("Groq_API_Key")
Llama3_8b = ChatGroq(temperature=0, model_name="Llama3-8b-8192")
Llama3_1_8b = ChatGroq(temperature=0, model_name="llama-3.1-8b-instant")
Llama3_2_13b = ChatGroq(temperature=0, model_name="llama-3.2-11b-text-preview")
Llama3_70b = ChatGroq(temperature=0.3, model_name="Llama3-70b-8192")
Llama3_70b_volatile = ChatGroq(temperature=1.0, model_name="Llama3-70b-8192")
Llama3_1_70b = ChatGroq(temperature=0.3, model_name="llama-3.1-70b-versatile")
GOOGLE_MAPS_API_KEY = os.getenv('GOOGLE_MAPS_API_KEY')
gmaps = googlemaps.Client(key=GOOGLE_MAPS_API_KEY)


class BaseDocumentData(BaseModel):
    """Base class for all document types with common fields"""
    extracted_date: str = Field(default_factory=str,description="Date when the information was extracted")
    confidence_score: float = Field(default_factory=str,description="Confidence score of extraction (0-1)")

class InsuranceApplicationData(BaseDocumentData):
    policy_holder_name: str = Field(default_factory=str,description="Full name of the policy holder")
    policy_holder_email: str = Field(default_factory=str,description="Email address of the policy holder")
    policy_type: str = Field(default_factory=str,description="Type of insurance policy")
    property_address: str = Field(default_factory=str,description="Complete address of the insured property")
    effective_date_of_coverage: str = Field(default_factory=str,description="Start date of insurance coverage")
    policy_holder_mobile: str = Field(default_factory=str,description="Mobile number of the policy holder")

class QuotesAndProposalsData(BaseDocumentData):
    Quote_or_Proposal_number: str = Field(default_factory=str,description="Quote or Proposal Number")
    Customer_name: str = Field(default_factory=str,description="Customer Full Name or Policy Holder Full Name")
    Effective_date: str = Field(default_factory=str,description="Effective Date of the Quote or Proposal")
    Type_of_coverage: str = Field(default_factory=str,description="Type of Coverage or Type of insurance policy")
    Agent_name: str = Field(default_factory=str,description="Agent Full Name")
    Underwriting_company_name: str = Field(default_factory=str,description="Underwriting Company Name or Insurance Company Name")

class PolicyDeclarationData(BaseDocumentData):
    Policy_number: str = Field(default_factory=str,description="Policy Number")
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Policy_effective_date: str = Field(default_factory=str,description="Policy Effective Date")
    Policy_expiration_date: str = Field(default_factory=str,description="Policy Expiration Date")
    Type_of_coverage: str = Field(default_factory=str,description="Type of Coverage or Type of insurance policy")
    Premium_amount: str = Field(default_factory=str,description="Premium Amount & add the $ sign to Premium Amount")  

class RenewalNoticeData(BaseDocumentData):
    Policy_number: str = Field(default_factory=str,description="Policy Number")
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Policy_expiration_date: str = Field(default_factory=str,description="Policy Expiration Date")
    Renewal_effective_date: str = Field(default_factory=str,description="Renewal Effective Date")
    Policy_type: str = Field(default_factory=str,description="Policy Type or Type of Coverage or Type of insurance policy")


class CancellationNoticeData(BaseDocumentData):
    Policy_number: str = Field(default_factory=str,description="Policy Number")
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Policy_effective_date: str = Field(default_factory=str,description="Policy Effective Date")
    Policy_expiration_date: str = Field(default_factory=str,description="Policy Expiration Date")
    Policy_cancellation_date: str = Field(default_factory=str,description="Policy Cancellation Date")
    Reason_for_policy_cancellation: str = Field(default_factory=str,description="Reason for Policy Cancellation")


class FirstNoticeOfLossData(BaseDocumentData): 
    policy_number: str = Field(default_factory=str,description="Policy Number")
    loss_date_and_time: str = Field(default_factory=str,description="Loss Date and Time")
    policy_effective_date: str = Field(default_factory=str,description="Policy Effective Date")
    policy_expiration_date: str = Field(default_factory=str,description="Policy Expiration Date")
    loss_location: str = Field(default_factory=str,description="Loss Location if not Loss Location consider Policy Holder Full Address")


class MedicalBillData(BaseDocumentData):
    Patient_name: str = Field(default_factory=str,description="Patient Full Name")
    Patient_ID: str = Field(default_factory=str,description="Patient ID")
    Guarantor_name: str = Field(default_factory=str,description="Guarantor Full Name")
    Guarantor_number: str = Field(default_factory=str,description="Guarantor Number")  
    Statement_date: str = Field(default_factory=str,description="Statement Date")


class CreditReportData(BaseDocumentData):
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Policy_holder_mobile: str = Field(default_factory=str,description="Policy Holder Mobile Number")
    Policy_holder_email: str = Field(default_factory=str,description="Policy Holder Email")
    Insurance_score: str = Field(default_factory=str,description="Insurance Score of the Policy Holder Insurance Company")
    Credit_inquiry_type: str = Field(default_factory=str,description="Credit Inquiry Type")   
    Credit_account_type: str = Field(default_factory=str,description="Credit Account Type")


class AppraisalReportData(BaseDocumentData):
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Policy_number: str = Field(default_factory=str,description="Policy Number")
    Claim_number: str = Field(default_factory=str,description="Claim Number")
    Appraiser_name: str = Field(default_factory=str,description="Appraiser Full Name")
    Appraisal_date: str = Field(default_factory=str,description="Appraisal Date")
    Date_of_loss: str = Field(default_factory=str,description="Date of Loss")
    Appraisal_id_number: str = Field(default_factory=str,description="Appraisal ID Number")

class InspectionReportData(BaseDocumentData):
    Policy_number: str = Field(default_factory=str,description="Policy Number")
    Policy_holder_name: str = Field(default_factory=str,description="Policy Holder Full Name")
    Property_address: str = Field(default_factory=str,description="Property Address")
    Inspection_date: str = Field(default_factory=str,description="Inspection Date")
    Inspection_type: str = Field(default_factory=str,description="Inspection Type")
    Inspection_id_number: str = Field(default_factory=str,description="Inspection ID Number")

class Sign_in_utils:
    @classmethod
    def generate_otp(cls):
        return ''.join(str(random.randint(0, 9)) for _ in range(6))

    @classmethod
    def send_verification_email(cls, email, verification_code):
        subject = 'Email Verification OTP'  # Updated subject
        context = {'verification_code': verification_code}
        html = os.path.join(settings.BASE_DIR, "templates", "email", "verification_email.html")
        message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER  # Use the configured email address
        recipient_list = [email]
        # Send HTML email with appropriate headers
        send_mail(
            subject,
            '',  # Empty message body as we're sending HTML
            'Innovon Technologies <' + email_from + '>',
            recipient_list,
            html_message=message,
        )


class Claim_utils:
    
    @classmethod
    def verify_policy_and_company(cls, policy_number, pol_date_of_birth=None):
        client, db = MongoDB.get_mongo_client_innoclaimfnol()
        policy_collection = db['policies']  # Assuming your collection name is 'Policy'
        company_collection = db['insurancecompanies']  # Assuming collection name is 'InsuranceCompany'

        try:
            policy = policy_collection.find_one({"policy_number": policy_number})
            if not policy:
                return None, None, 'No user exists with this policy number'
            if pol_date_of_birth != None:
                # Date formatting and comparison (adjust timezone if needed)
                original_date = datetime.strptime(pol_date_of_birth, "%d/%m/%Y").date()
                offset = timedelta(hours=5, minutes=30)  # Assuming UTC+5:30
                formatted_date = (original_date + offset).isoformat()

            if pol_date_of_birth != None and policy['pol_date_of_birth'] != formatted_date:
                return None, None, 'The date of birth does not match with the records'

            company = company_collection.find_one({"ic_id": policy['ic_id']})
            if not company:
                return None, None, 'Company information not found'

            policy = Claim_utils.convert_objectid_to_str(policy)
            company = Claim_utils.convert_objectid_to_str(company)
            return policy, company, 'Verification successful'

        except Exception as error:
            return None, None, str(error)
        
    
    @classmethod
    def convert_objectid_to_str(cls, data):
        if isinstance(data, ObjectId):
            return str(data)
        elif isinstance(data, dict):
            return {key: Claim_utils.convert_objectid_to_str(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [Claim_utils.convert_objectid_to_str(element) for element in data]
        else:
            return data
        
    @classmethod
    def generate_next_claim_id(cls, latest_claim_id):
        current_year = datetime.today().strftime("%Y")  # Get current year as string

        if latest_claim_id:
            prefix = latest_claim_id[:3]  # Extract prefix (e.g., "CLM")
            numeric_part = int(latest_claim_id[7:])  # Extract numeric part
        else:
            prefix = "CLM"  # Default prefix
            numeric_part = 1000  # Starting number

        next_numeric_part = numeric_part + 1 if numeric_part < 999999 else 1
        next_claim_id = f"{prefix}{current_year}{str(next_numeric_part).zfill(6)}"
        return next_claim_id
    
    @classmethod
    def generate_next_ic_id(cls, latest_ic_id):
        print(latest_ic_id)
        prefix = latest_ic_id[:2] if latest_ic_id else "IC"
        print("prefix: ", prefix)
        numeric_part = int(latest_ic_id[2:]) if latest_ic_id else 0
        print("numeric_part: ", numeric_part)
        next_numeric_part = numeric_part + 1 if numeric_part < 999999 else 1
        return f"{prefix}{str(next_numeric_part).zfill(6)}"
    
    @classmethod
    def get_dictionary_structure(cls, extracted_data, role, policy=None, email=None):
        claimants_contact = []
        if role != "Insured":
            claimant_data = {
                "role": role,
                "contact": {
                    "first_name": extracted_data.get('First_Name', "Null"),
                    "middle_name": extracted_data.get('Middle_Name', "Null"),
                    "last_name": extracted_data.get('Last_Name', "Null"),
                    "phone_number": extracted_data.get('Mobile_Number', "Null"),
                    "email": email,
                    "relationship_with_insured": extracted_data.get('relationship_with_insured', "Null"),
                    "street_number": extracted_data.get('Claimant_street_number', "Null"),
                    "street_name": extracted_data.get('Claimant_street_name', "Null"),
                    "city": extracted_data.get('Claimant_city', "Null"),
                    "zip": extracted_data.get('Claimant_zip', "Null"),
                    "state": extracted_data.get('Claimant_state', "Null"),
                    "country": extracted_data.get('Claimant_country', "Null"),
                    "proof_of_identity": extracted_data.get('Proof_of_Identity', "Null"),
                    "proof_of_identity_number": extracted_data.get('Proof_of_Identity_Number', "Null"),
                    "social_security_number": extracted_data.get('Social_Security_Number', "Null")
                }
            }
        else:
            claimant_data = {
                "role": role,
                "contact": {
                    "first_name": policy.get('pol_first_name', "Null"),
                    "middle_name": policy.get('pol_middle_name', "Null"),
                    "last_name": policy.get('pol_last_name', "Null"),
                    "phone_number": policy.get('pol_mobile', "Null"),
                    "email": policy.get('pol_email', "Null"),
                    "street_number": policy.get('pro_address1', "Null"),
                    "street_name": policy.get('pro_street', "Null"),
                    "city": policy.get('pro_city', "Null"),
                    "zip": policy.get('pro_zip', "Null"),
                    "state": policy.get('pro_state', "Null"),
                    "country": policy.get('pro_country', "Null"),
                }
            }
        claimants_contact.append(claimant_data)
        # Convert to the desired format
        converted_data = {
            "policy_number": extracted_data.get('policy_number'),
            "loss_date_and_time": extracted_data.get('loss_date_and_time'),
            "loss_type": extracted_data.get('loss_type'),
            "loss_property": extracted_data.get('loss_property'),
            "loss_damage_description": extracted_data.get('loss_damage_description'),
            "street_number": extracted_data.get('street_number'),
            "street_name": extracted_data.get('street_name'),
            "loss_city": extracted_data.get('loss_city'),
            "loss_state": extracted_data.get('loss_state'),
            "loss_zip": extracted_data.get('loss_zip'),
            "loss_country": extracted_data.get('loss_country'),
            "police_fire_contacted": extracted_data.get('police_fire_contacted'),
            "report_number": extracted_data.get('report_number'),
            "claim_reported_by": role,
            "claimants_contact": claimants_contact,
            "claim_process_document_name": extracted_data.get('claim_process_document_name', "Null"),
            "claim_process_document_url": extracted_data.get('claim_process_document_url', "Null"),
            "claim_witness_document_names": extracted_data.get('claim_witness_document_names', "Null"),
            "claim_witness_document_urls": extracted_data.get('claim_witness_document_urls', "Null"),
            "claim_created_at": extracted_data.get('claim_created_at', None),
            "claim_id": extracted_data.get('claim_id')
        }
        return converted_data
    
    @classmethod
    def get_dictionary_structure_for_documents(cls, extracted_data, role, user_email=None):
        claimants_contact_str = ""
        if role != "Insured":
            claimants_contact = [extracted_data.get('First_Name', "Null"), extracted_data.get('Middle_Name', "Null"),
                                extracted_data.get('Last_Name', "Null"), str(extracted_data.get('Mobile_Number', "Null")),
                                user_email, extracted_data.get('relationship_with_insured', "Null"),
                                str(extracted_data.get('Claimant_street_number', "Null")),
                                extracted_data.get('Claimant_street_name', "Null"),
                                extracted_data.get('Claimant_city', "Null"),
                                str(extracted_data.get('Claimant_zip', "Null")),
                                extracted_data.get('Claimant_state', "Null"),
                                extracted_data.get('Claimant_country', "Null"),
                                extracted_data.get('Proof_of_Identity', "Null"),
                                str(extracted_data.get('Proof_of_Identity_Number', "Null")),
                                str(extracted_data.get('Social_Security_Number', "Null"))]
            claimants_contact_str = ",".join(claimants_contact)

        converted_data = {
            "policy_number": extracted_data.get('policy_number'),
            "loss_date_and_time": extracted_data.get('loss_date_and_time'),
            "loss_type": extracted_data.get('loss_type'),
            "loss_property": extracted_data.get('loss_property'),
            "loss_damage_description": extracted_data.get('loss_damage_description'),
            "street_number": extracted_data.get('street_number'),
            "street_name": extracted_data.get('street_name'),
            "loss_city": extracted_data.get('loss_city'),
            "loss_state": extracted_data.get('loss_state'),
            "loss_zip": extracted_data.get('loss_zip'),
            "loss_country": extracted_data.get('loss_country'),
            "police_fire_contacted": extracted_data.get('police_fire_contacted'),
            "report_number": extracted_data.get('report_number'),
            "claim_reported_by": role,
            "insured_contact_details": extracted_data.get('insured_contact_details', "Null"),
            "non_insured_contact_details": extracted_data.get('non_insured_contact_details', "Null"),
            "claim_process_document_name": extracted_data.get('claim_process_document_name', "Null"),
            "claim_process_document_url": extracted_data.get('claim_process_document_url', "Null"),
            "claim_witness_document_names": extracted_data.get('claim_witness_document_names', "Null"),
            "claim_witness_document_urls": extracted_data.get('claim_witness_document_urls', "Null"),
            "claim_created_at": extracted_data.get('claim_created_at'),
            "claim_id": extracted_data.get('claim_id')
        }
        if role != "Insured":
            converted_data['non_insured_contact_details'] = claimants_contact_str
        return converted_data
    
    @classmethod
    def add_claim_to_db(cls, claim_data):
        if 'loss_type' not in claim_data:
            claim_data['loss_type'] = 'fire'
        if 'claim_reported_by' not in claim_data:
            claim_data['claim_reported_by'] = 'Insured'
        if 'claim_witness_document_names' not in claim_data:
            claim_data['claim_witness_document_names'] = 'None'
        if 'claim_witness_document_urls' not in claim_data:
            claim_data['claim_witness_document_urls'] = 'None'
        if 'claim_process_document_name' not in claim_data:
            claim_data['claim_process_document_name'] = 'None'
        if 'claim_process_document_url' not in claim_data:
            claim_data['claim_process_document_url'] = 'None'

        # Datetime conversion (if loss_date_and_time is a string)
        if isinstance(claim_data['loss_date_and_time'], str):
            try:
                # Parse the date and time using the frontend's format
                dt_object_loss = datetime.strptime(claim_data['loss_date_and_time'], "%Y/%m/%d %H:%M:%S")
                # Convert to ISO 8601 format with timezone information
                claim_data['loss_date_and_time'] = dt_object_loss.isoformat()
            except ValueError:
                raise ValueError("Invalid format for loss_date_and_time. Use YYYY/MM/DD HH:MM:SS format.")

        client, db = MongoDB.get_mongo_client_innoclaimfnol()
        claim_collection = db['claims']
        encrypted_data = Authentication.encrypt_data(claim_data)
        inserted_claim = claim_collection.insert_one(encrypted_data)
        return inserted_claim
    
    @classmethod
    def add_company_to_db(cls, company_data):
        # Required fields
        required_fields = [
            'ic_id', 'ic_name', 'ic_address1', 'ic_street', 'ic_city',
            'ic_zip', 'ic_state', 'ic_country', 'ic_mobile', 'ic_email',
            'ic_primary_color', 'ic_website_url',
            # 'ic_logo_name', 'ic_logo_path'
        ]

        # Check for missing required fields
        for field in required_fields:
            if field not in company_data or not company_data[field]:
                raise ValueError(f"Missing required field: {field}")

        if "ic_address2" not in company_data:
            company_data["ic_address2"] = "null"

        if "ic_secondary_color" not in company_data:
            company_data["ic_secondary_color"] = "null"

        if not 10 <= len(str(company_data['ic_mobile'])) <= 11:
            raise ValueError("Mobile number must contain only 10-11 digits, Please re-verify and enter")

        # ... (Additional validations as needed)

        client, db = MongoDB.get_mongo_client_innoclaimfnol()
        company_collection = db['insurancecompanies']  # Assuming collection name
        inserted_company = company_collection.insert_one(company_data)
        return inserted_company
    
    @classmethod
    def get_user_email_from_policy(cls, policy_number):
        client, db = MongoDB.get_mongo_client_innoclaimfnol()
        policy_collection = db['policies']
        user_policy_data = policy_collection.find_one({"policy_number": policy_number})
        if user_policy_data:
            return user_policy_data['pol_email']
        else:
            return None  # Or raise an exception
    
    @classmethod
    def get_policy_numbers_from_db(cls):
        """Retrieves all policy numbers from the MongoDB 'policies' collection."""

        client, db = MongoDB.get_mongo_client_innoclaimfnol()
        policies = db['policies'].find({}, {'policy_number': 1, '_id': 0})  # Only fetch policy_number field
        policy_numbers = [policy['policy_number'] for policy in policies]
        return policy_numbers
            
    
        


class Address_validations:

    @classmethod
    def extract_component(cls, components, component_type):
        """Helper function to extract a specific address component."""
        for component in components:
            if component_type in component['types']:
                return component['long_name']
        return None

    @classmethod
    def reverse_geocode(cls, api_key, latitude, longitude):
        base_url = "https://maps.googleapis.com/maps/api/geocode/json"
        endpoint = f"{base_url}?latlng={latitude},{longitude}&key={api_key}"

        try:
            response = requests.get(endpoint)
            response.raise_for_status()
            data = response.json()

            if data['status'] == 'OK':
                address_components = data['results'][0]['address_components']
                address = {
                    'street_number': Address_validations.extract_component(address_components, 'street_number'),
                    'street_name': Address_validations.extract_component(address_components, 'route'),
                    'city': Address_validations.extract_component(address_components, 'locality'),
                    'state': Address_validations.extract_component(address_components, 'administrative_area_level_1'),
                    'zipcode': Address_validations.extract_component(address_components, 'postal_code'),
                    'country': Address_validations.extract_component(address_components, 'country')
                }
                return address
            else:
                print(f"Geocoding API request failed with status: {data['status']}")
                return None

        except requests.exceptions.RequestException as e:
            print(f"An error occurred during the request: {e}")
            return None
        
    @classmethod
    def parse_address(cls, address_string):
        try:
            parsed_address = usaddress.tag(address_string)[0]
            address_parts = {
                'street_number': int(parsed_address.get('AddressNumber',0)) if parsed_address.get('AddressNumber', None) is not None else None,
                'street_name': parsed_address.get('StreetName', None),
                'city': parsed_address.get('PlaceName', None),
                'state': parsed_address.get('StateName', None),
                'zip_code': int(parsed_address.get('ZipCode', 0))if parsed_address.get('ZipCode', None) is not None else None,
                'country': "USA"
            }
            return address_parts
        except usaddress.RepeatedLabelError:
            return {
                'street_number': None,
                'street_name': None,
                'city': None,
                'state': None,
                'zip_code': None,
                'country': None
            }
    
    # @classmethod
    # def validate_address(cls, address_string):
    #     result = gmaps.places(query=address_string)
    #     formatted_address = None
    #     if result['results']:
    #         # Extract the formatted address
    #         formatted_address = result['results'][0]['formatted_address']
    #     if formatted_address:
    #         return formatted_address
    #     else:
    #         return "Address Not validated"
    @classmethod
    def validate_address(cls, address_string):
        if address_string=='':
            return "Address Not validated"
        else:
            result = gmaps.places(query=address_string)
            formatted_address = None           
            if result['results']:
                place_details = result['results'][0]
                if 'place_id' in place_details and 'geometry' in place_details:
                    formatted_address = place_details.get('formatted_address')
                    if formatted_address and ',' in formatted_address:   
                        return formatted_address           
            return "Address Not validated"
        
    @classmethod
    def format_address(cls, info, prefix):
        fields = [
            f"{prefix}_street_number",
            f"{prefix}_street_name",
            f"{prefix}_city",
            f"{prefix}_state",
            f"{prefix}_zip",
            f"{prefix}_country"
        ]
        return " ".join(str(info.get(field, '')) for field in fields).strip()
        

class File_handling:

    @classmethod
    def save_documents_with_unique_ids(cls, documents, save_directory):
        original_filenames = []
        document_urls = []  # Store URLs instead of paths
        for document in documents:
            # Generate unique ID
            unique_id = str(uuid.uuid4())
            # Get file extension
            original_filename, original_extension = os.path.splitext(document.name)
            # Create new filename with unique ID
            new_filename = f"{unique_id}{original_extension}"
            # Construct full save path
            save_path = os.path.join(settings.MEDIA_ROOT, save_directory, new_filename)
            # Save the file
            with open(save_path, 'wb+') as destination:
                for chunk in document.chunks():
                    destination.write(chunk)
            # Store filename and URL
            original_filenames.append(document.name)
            document_urls.append(f"{settings.MEDIA_URL}{save_directory}/{new_filename}")
        return original_filenames, document_urls, "Files saved successfully"
    
    @classmethod
    def save_policy_documents_with_unique_ids(cls, document, save_directory):
        document_url = ''  # Store URLs instead of paths
        unique_id = str(uuid.uuid4())
        # Get file extension
        original_filename, original_extension = os.path.splitext(document.name)
        # Create new filename with unique ID
        new_filename = f"{unique_id}{original_extension}"
        # Construct full save path
        save_path = os.path.join(settings.MEDIA_ROOT, save_directory, new_filename)
        # Save the file
        with open(save_path, 'wb+') as destination:
            for chunk in document.chunks():
                destination.write(chunk)
        # Store filename and URL
        document_url = f"{settings.MEDIA_URL}{save_directory}/{new_filename}"
        return original_filename, document_url, "Files saved successfully"
    
    @classmethod
    def save_to_excel(cls, claim_data, company_name):
        """Saves claim data to an Excel file, appending if the file exists."""
        EXCEL_SHEETS_DIR = os.path.join(settings.MEDIA_ROOT, 'Excel_sheets')

        file_path = os.path.join(EXCEL_SHEETS_DIR, f"{company_name}.xlsx")

        try:
            next_claim_id = File_handling.generate_next_claim_id_for_documents(file_path, "excel")
            claim_data['claim_id'] = next_claim_id  # Add the new claim_id to your claim_data
            claim_data = File_handling.convert_datetime_to_string(claim_data)

            existing_headers = []
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                existing_headers = [cell.value for cell in ws[1]]
                wb.close()
                # 2. Structure the row using existing headers
                structured_row = {}
                for header in existing_headers:
                    structured_row[header] = claim_data.get(header, '')
            if os.path.exists(file_path):
                df = pd.DataFrame([structured_row])
                wb = load_workbook(file_path)
                ws = wb.active
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
            else:
                df = pd.DataFrame([claim_data])
                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            wb.save(file_path)
            wb.close()
            message = f"Claim data saved to {file_path}"
            return message, next_claim_id

        except Exception as e:
            return f"Error saving to Excel: {e}", None

    @classmethod
    def save_to_csv(cls, claim_data, company_name):
        """Saves claim data to a CSV file, appending if the file exists."""
        CSV_FILES_DIR = os.path.join(settings.MEDIA_ROOT, 'csv_files')
        file_path = os.path.join(CSV_FILES_DIR, f"{company_name}.csv")

        try:
            next_claim_id = File_handling.generate_next_claim_id_for_documents(file_path, "csv")
            claim_data['claim_id'] = next_claim_id

            # 1. Read existing headers
            existing_headers = []
            if os.path.exists(file_path):
                with open(file_path, 'r') as f:
                    existing_headers = f.readline().strip().split(',')  # Read and split the first line

            # 2. Structure the row
            structured_row = {}
            for header in existing_headers:
                structured_row[header] = claim_data.get(header, '')

            # 4. Append data to the CSV
            if os.path.exists(file_path):
                # 3. Create DataFrame
                df = pd.DataFrame([structured_row])
                df.to_csv(file_path, mode='a', header=False, index=False)
            else:
                df = pd.DataFrame([claim_data])
                df.to_csv(file_path, mode='w', header=True, index=False)

            message = f"Claim data saved to {file_path}"
            return message, next_claim_id

        except Exception as e:
            return f"Error saving to CSV: {e}", None

    @classmethod
    def save_to_flat_file(cls, claim_data, company_name):
        """Saves claim data to a fixed-width flat file (.txt)."""

        FLAT_FILES_DIR = os.path.join(settings.MEDIA_ROOT, 'flat_files')
        os.makedirs(FLAT_FILES_DIR, exist_ok=True)  # Create directory if it doesn't exist

        file_path = os.path.join(FLAT_FILES_DIR, f"{company_name}.txt")

        # Define field widths (adjust as needed based on your requirements)
        field_widths = {
            'policy_number': 10,
            'loss_date_and_time': 19,
            'loss_type': 30,
            'loss_property': 221,
            'loss_damage_description': 500,
            'street_number': 12,
            'street_name': 80,
            'loss_city': 50,
            'loss_state': 45,
            'loss_zip': 10,
            'loss_country': 24,
            'police_fire_contacted': 4,
            'report_number': 12,
            'claim_reported_by': 30,
            'claim_storage_type': 10,
            'insured_contact_details': 380,
            'non_insured_contact_details': 440,
            'claim_witness_document_names': 75,
            'claim_witness_document_urls': 75,
            'claim_process_document_name': 75,
            'claim_process_document_url': 75,
            'claim_created_at': 35,
            'claim_id': 13,
        }

        try:
            # Generate the next claim_id
            next_claim_id = File_handling.generate_next_claim_id_for_documents(file_path, "txt")
            claim_data['claim_id'] = next_claim_id

            # Format the data for fixed-width output
            formatted_data = ''
            for field, width in field_widths.items():
                value = str(claim_data.get(field, '')).ljust(width)  # Get value or '' if missing
                formatted_data += value

            # Append the formatted data to the file
            with open(file_path, 'a') as f:
                f.write(formatted_data + '\n')

            message = f"Claim data saved to {file_path}"
            return message, next_claim_id

        except Exception as e:
            return f"Error saving to flat file: {e}", None
        
    @classmethod
    def generate_next_claim_id_for_documents(cls, file_path, file_type):
        """
        Generates the next claim ID based on existing data in the file.

        Args:
            file_path (str): The full path to the Excel or CSV file.
            file_type (str): The type of file ("excel" or "csv").

        Returns:
            str: The next claim ID in the format "CLM<year><6-digit number>".
        """
        current_year = datetime.today().strftime("%Y")
        latest_claim_id = None

        if os.path.exists(file_path):
            if file_type == "excel":
                try:
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active
                    for row in reversed(list(ws.iter_rows(values_only=True))):
                        if row[-1] is not None:  # Assuming claim_id is in the first column
                            latest_claim_id = str(row[-1])
                            break
                    wb.close()
                except Exception as e:
                    print(f"Error reading Excel file: {e}")
            elif file_type == "csv":
                try:
                    with open(file_path, 'r') as f:
                        lines = f.readlines()
                        if lines:
                            last_line = lines[-1]
                            latest_claim_id = last_line.split(',')[-1].strip()  # Assuming CSV uses comma as delimiter
                except Exception as e:
                    print(f"Error reading CSV file: {e}")
            elif file_type == "txt":  # For Flat Files
                try:
                    with open(file_path, 'r') as f:
                        for line in f:  # Read line by line (important for large files)
                            latest_claim_id = line[-14:].strip()  # Extract the last 13 characters (claim_id)
                        # if latest_claim_id:  # Check if a claim_id was found
                        #     break
                except Exception as e:
                    print(f"Error reading flat file: {e}")

        if latest_claim_id:
            prefix = latest_claim_id[:3]
            numeric_part = int(latest_claim_id[7:])
        else:
            prefix = "CLM"
            numeric_part = 1000

        next_numeric_part = numeric_part + 1 if numeric_part < 999999 else 1
        next_claim_id = f"{prefix}{current_year}{str(next_numeric_part).zfill(6)}"
        return next_claim_id

    @classmethod
    def convert_datetime_to_string(cls, data):
        """Converts datetime objects in the data to strings."""
        value = data['claim_created_at']
        data['claim_created_at'] = value.strftime('%Y-%m-%d %H:%M:%S')
        return data
    
    @classmethod
    def extract_text_from_pdf(cls, pdf_file_object):
        """
        Extracts form data and text from the first two pages of a PDF.

        Args:
            pdf_file_object: The PDF file object.

        Returns:
            Extracted text from form data and text from the first two pages of a PDF into a single string.
        """
        # Open the PDF file
        doc = fitz.open(stream=pdf_file_object, filetype="pdf")

        # Dictionary to store all extracted information
        pdf_data = {
            'text_content': [],
            'form_fields': [],
            'checkboxes': []
        }

        # Iterate through each page
        for page_num in range(len(doc)):
            page = doc[page_num]

            # Extract text content
            text = page.get_text()
            pdf_data['text_content'].append(f"Page {page_num + 1}:\n{text}\n")

            # Extract form fields and checkboxes
            widgets = page.widgets()
            for widget in widgets:
                if widget.field_type == fitz.PDF_WIDGET_TYPE_TEXT:
                    pdf_data['form_fields'].append({
                        'page': page_num + 1,
                        'name': widget.field_name,
                        'value': widget.field_value,
                    })
                elif widget.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                    pdf_data['checkboxes'].append({
                        'page': page_num + 1,
                        'name': widget.field_name,
                        'checked': widget.field_value,
                    })

        # Close the document
        doc.close()
        combined_len = pdf_data['text_content'] + pdf_data['form_fields'] + pdf_data['checkboxes']

        if len(combined_len[0]) < 10:
            return "pdf_image"
        return pdf_data
    
    @classmethod
    def extract_text_from_docx(cls, docx_file_object):
        """Extracts text from a DOCX file object."""
        doc = docx.Document(docx_file_object)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text

    @classmethod
    def extract_text_from_txt(cls, txt_file_object):
        """Extracts text from a TXT file object."""
        text = txt_file_object.read().decode('utf-8')
        return text
    
    @classmethod
    def extract_images_from_pdf(cls, pdf_bytes):
        """Processes a PDF by extracting all pages as images and returning them as bytes."""
        if not pdf_bytes.getbuffer().nbytes:
            raise ValueError("The PDF file is empty or not readable.")

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")  # Open PDF from bytes
        image_bytes_list = []

        for page_index in range(len(doc)):
            page = doc[page_index]
            # Zoom if necessary (adjust zoom_x and zoom_y as needed)
            zoom_x = 3.0
            zoom_y = 3.0
            mat = fitz.Matrix(zoom_x, zoom_y)
            pix = page.get_pixmap(matrix=mat)
            # Convert pixmap to bytes
            image_bytes = pix.tobytes("png")
            image_bytes_list.append(image_bytes)

        return image_bytes_list



class Emails:

    @classmethod
    def send_claim_confirmation_email(cls, email, claim_id, company_name, insured_name, company_logo, company_email,
                                  company_phone, company_color, ic_id):
        path = os.path.join(settings.BASE_DIR, 'templates', 'Assets')
        subject = 'Claim Submission Confirmation'
        company_color = str(company_color)[1:]
        context = {
            'insured_email': email,
            'insured_name': insured_name,
            'claim_id': claim_id,
            'company_name': company_name,
            'company_logo': company_logo,
            'company_email': company_email,
            'company_phone': company_phone,
            'company_color': company_color,
            'ic_id': ic_id,
            'path': path
        }
        html = os.path.join(settings.BASE_DIR, "templates", "email", "claim_confirmation_email.html")
        message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [email]
        send_mail(
            subject,
            '',
            f'{company_name} Claims <{email_from}>',  # Updated "From" address 
            recipient_list,
            html_message=message,
        )
    
    @classmethod
    def send_claim_failure_email(cls, email, missing_fields):
        subject = 'Important: Your Claim Submission Requires Attention'

        context = {
            'missing_fields': missing_fields,
            'support_email': 'info@innovontek.com',
            'support_phone': '8008673672',
            'company_name': 'Innovon Technologies'
        }
        html = os.path.join(settings.BASE_DIR, "templates", "email", "claim_submission_failed.html")
        html_message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER
        plain_message = strip_tags(html_message)

        from_email = f"{'Innovon Technologies'} Claims <{email_from}>"
        recipient_list = [email]

        try:
            send_mail(
                subject,
                plain_message,
                from_email,
                recipient_list,
                html_message=html_message,
                fail_silently=False,
            )
            print(f"Claim failure email sent successfully to {email}")
        except Exception as e:
            print(f"Failed to send claim failure email to {email}. Error: {str(e)}")
        return

    @classmethod
    def send_claim_failure_email_missing(cls, email, missing_fields, company_name, insured_name, company_logo, company_email,
                                        company_phone, company_color, ic_id):

        subject = 'Claim Submission Confirmation'
        context = {
            'insured_email': email,
            'insured_name': insured_name,
            'missing_values': missing_fields,
            'company_name': company_name,
            'company_logo': company_logo,
            'company_email': company_email,
            'company_phone': company_phone,
            'company_color': company_color,
            'ic_id': ic_id,
        }
        html = os.path.join(settings.BASE_DIR, "templates", "email", "claim_failure_confirmation_email.html")
        message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [email]
        send_mail(
            subject,
            '',
            f'{company_name} Claims <{email_from}>',  # Updated "From" address
            recipient_list,
            html_message=message,
        )
    
    @classmethod
    def send_policy_intake_success_mail(cls, policy_number, name, policy_type, email):
        subject = 'Policy Submission Confirmation'

        # Read and encode the logo
        logo_path = os.path.join(settings.MEDIA_ROOT, 'ITlogo.png')
        with open(logo_path, "rb") as image_file:
            encoded_logo = base64.b64encode(image_file.read()).decode()
        context = {
            'policy_number': policy_number,
            'policy_type': policy_type,
            'name': name,
            'company_logo': encoded_logo,  # Add the encoded logo to the context
            'support_email': 'info@innovontek.com',
            'support_phone': '8008673672',
            'company_name': 'Innovon Technologies',
            'company_website': 'https://innovontek.com/'
        }
        html = os.path.join(settings.BASE_DIR, "templates", "email", "policy_intake_confirmation.html")
        message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [email]
        try:
            send_mail(
                subject,
                strip_tags(message),  # Plain text alternative for some email clients
                'Innovon Technologies <' + email_from + '>',
                recipient_list,
                html_message=message,
                fail_silently=False,
            )
            print(f"Policy Success email sent successfully to {email}")
        except Exception as e:
            print(f"Failed to send claim failure email to {email}. Error: {str(e)}")
        return
    
    @classmethod
    def send_demo_alert_email(cls, details):
        """
        Sends a demo alert email to the CEO with customer details and company logo.
        """
        subject = 'New Demo Request - Innovon Technologies'

        # Read and encode the logo
        logo_path = os.path.join(settings.MEDIA_ROOT, 'ITlogo.png')
        with open(logo_path, "rb") as image_file:
            encoded_logo = base64.b64encode(image_file.read()).decode()

        context = {
            'user_details': details,
            'company_color': '#010066',  # Company color code
            'accent_color': '#0B70FF',  # Accent color code
            'company_logo': encoded_logo  # Add the encoded logo to the context
        }

        html = os.path.join(settings.BASE_DIR, "templates", "email", "demo_alert_email.html")
        message = render_to_string(html, context)
        email_from = settings.EMAIL_HOST_USER
        recipient_list = ['info@innovontek.com']  # CEO's email

        send_mail(
            subject,
            strip_tags(message),  # Plain text alternative for some email clients
            'Innovon Technologies <' + email_from + '>',
            recipient_list,
            html_message=message,
            fail_silently=False,
        )


class PolicyInfo(BaseModel):
        selectedPolicy: str = Field(default_factory=str)
        policy_holder_FirstName: str = Field(default_factory=str)
        policy_holder_LastName: str = Field(default_factory=str)
        policy_holder_street_number: int = Field(default_factory=int)
        policy_holder_street_name: str = Field(default_factory=str)
        policy_holder_city: str = Field(default_factory=str)
        policy_holder_state: str = Field(default_factory=str)
        policy_holder_zip: int = Field(default_factory=int)
        policy_holder_country: str = Field(default_factory=str)
        policy_holder_mobile: str = Field(default_factory=str)
        policy_holder_email: str = Field(default_factory=str)
        policy_holder_occupation: str = Field(default_factory=str)
        policy_holder_ssn: str = Field(default_factory=str)


class PropertyInfo(BaseModel):
    residenceType: str = Field(default_factory=str)
    constructionType: str = Field(default_factory=str)
    yearBuilt: int = Field(default_factory=int)
    numberOfStories: int = Field(default_factory=int)
    squareFootage: int = Field(default_factory=int)
    heatingType: str = Field(default_factory=str)
    plumbing_installed_year: int = Field(default_factory=int)
    wiring_installed_year: int = Field(default_factory=int)
    heating_installed_year: int = Field(default_factory=int)
    roof_installed_year: int = Field(default_factory=int)
    fireHydrantDistance: float = Field(default_factory=float)
    fireStationDistance: float = Field(default_factory=float)
    alternateHeating: str = Field(default_factory=str)
    any_business_conducted_on_premises: str = Field(default_factory=str)
    trampolineRamp: str = Field(default_factory=str)
    subjectToFlood: str = Field(default_factory=str)
    floodInsuranceRequested: str = Field(default_factory=str)
    rentedToOthers: str = Field(default_factory=str)
    CoverageLocation_street_number: int = Field(default_factory=int)
    CoverageLocation_street_name: str = Field(default_factory=str)
    CoverageLocation_city: str = Field(default_factory=str)
    CoverageLocation_state: str = Field(default_factory=str)
    CoverageLocation_zip: int = Field(default_factory=int)
    CoverageLocation_country: str = Field(default_factory=str)
    additionalInfo: Optional[str] = Field(default_factory=str)


class AdditionalInfo(BaseModel):
    currentInsuranceCarrier: str = Field(default_factory=str)
    currentPolicy: str = Field(default_factory=str)
    effectiveDate: datetime = Field(..., description="Date in YYYY-MM-DD format")
    current_policy_premium: float = Field(default_factory=float)
    anyLossLast4Years: str = Field(default_factory=str)
    mortgageeName: Optional[str] = Field(default_factory=str)
    mortgageeStreetNumber: Optional[int] = Field(default_factory=int)
    mortgageeStreetName: Optional[str] = Field(default_factory=str)
    mortgageeCity: Optional[str] = Field(default_factory=str)
    mortgageeState: Optional[str] = Field(default_factory=str)
    mortgageeCountry: Optional[str] = Field(default_factory=str)
    mortgageeZip: Optional[int] = Field(default_factory=int)
    mortgageeInstallmentAmount: Optional[float] = Field(default_factory=float)


class Coverages(BaseModel):
    dwellingCoverage: float = Field(default_factory=float)
    personalProperty: float = Field(default_factory=float)
    personalLiabilityCoverage: float = Field(default_factory=float)
    medicalPayments: float = Field(default_factory=float)
    deductible: float = Field(default_factory=float)


class patient_info(BaseModel):
    patient_name: str = Field(default_factory=str)
    patient_address: str = Field(default_factory=str)
    account_number: str = Field(default_factory=str)


class guarantor_info(BaseModel):
    guarantor_name: str = Field(default_factory=str)
    guarantor_address: str = Field(default_factory=str)
    guarantor_number: str = Field(default_factory=str)


class service_info(BaseModel):
    hospital_address: str = Field(default_factory=str)
    service_doctor: str = Field(default_factory=str,description="The name of the doctor who provided the service to the patient")
    statement_date: str = Field(default_factory=str)
    date_of_service: str = Field(default_factory=str)
    charges: str = Field(default_factory=str, description="Along with the dollar '$' symbol")
    payments_adjustments: str = Field(default_factory=str,description="The payments and adjustments along with the '$' symbol")
    insurance_payments_adjustments: str = Field(default_factory=str)
    patient_balance: str = Field(default_factory=str, description="patient balance, Along with the dollar '$' symbol")
    billing_support_contact: int = Field(default_factory=int)


class Ai_utils:

    @staticmethod
    def returns_token_count(response):
        output_tokens = response.response_metadata['token_usage']['completion_tokens']
        input_tokens = response.response_metadata['token_usage']['prompt_tokens']
        total_tokens = response.response_metadata['token_usage']['total_tokens']
        return output_tokens, input_tokens, total_tokens
    
    @staticmethod
    def extract_data_from_error(error_message: str) -> Dict:
        """Extracts key-value pairs from the error message.

        Args:
            error_message (str): The error message string.

        Returns:
            Dict: A dictionary containing the extracted key-value pairs, or an empty dictionary if no data is found.
        """

        pattern = r'"(.*?)"\s*:\s*(.*?)\s*(?:\/\/.*?)?,?\s*$'
        matches = re.findall(pattern, error_message, re.MULTILINE)

        extracted_data = {}
        for match in matches:
            key, value = match
            value = value.strip()

            if value.lower() == "null":
                value = None
            elif value.lower() == "true":
                value = True
            elif value.lower() == "false":
                value = False
            else:
                value = value.strip('"')

            extracted_data[key] = value

        # Clean up extra comma in loss_date_and_time (if present)
        if '",' in extracted_data.get('loss_date_and_time', ''):
            extracted_data['loss_date_and_time'] = extracted_data['loss_date_and_time'][:-2]

        return extracted_data

    @classmethod
    def get_textract_queries_for_claim(cls):
        queries = [
            {'Text': 'What is policy number?', 'Alias': 'policy_number'},
            {'Text': ' Who reported the claim ?', 'Alias': 'claim_reported_by'},
            {'Text': 'what is the loss date and time?', 'Alias': 'loss_date_and_time'},
            {'Text': 'What is the cause for loss(FIRE/STORM/...)?', 'Alias': 'loss_type'},
            {'Text': 'What is the loss street number', 'Alias': 'street_number'},
            {'Text': 'What is the loss street name ?', 'Alias': 'street_name'},
            {'Text': 'What is the loss city ?', 'Alias': 'loss_city'},
            {'Text': 'What is the loss STATE name ?', 'Alias': 'loss_state'},
            {'Text': 'What is the loss zip ?', 'Alias': 'loss_zip'},
            {'Text': 'What is the loss country ?', 'Alias': 'loss_country'},
            {'Text': 'is there any police or fire contacted checked?', 'Alias': 'police_fire_contacted'},
            {'Text': 'what is the report number of police or fire department contacted?', 'Alias': 'report_number'},
            {'Text': 'What was happened to the insured home?', 'Alias': 'loss_damage_description'},
        ]
        return queries
    
    @classmethod
    def add_LLM_Metadata_to_database(cls, input_tokens, output_tokens, total_tokens, accuracy, reason, filename, application,):
        client, db = MongoDB.get_mongo_client_Administration()
        llm_metadata_collection = db['llm_metadata']
        meta_data = {'file_name': filename,
                    'application': application,
                    'accuracy': accuracy,
                    'reason': reason,
                    'input_tokens': input_tokens,
                    'output_tokens': output_tokens,
                    'total_tokens': total_tokens}
        try:
            result = llm_metadata_collection.insert_one(meta_data)
            return "Data Added Successfully"
        except Exception as e:
            return str(e)
        
    @classmethod
    def extract_text_from_image(cls, image_file_object, queries):
        """Extracts text from an image using Amazon Textract."""

        textract_client = boto3.client('textract',
                                    aws_access_key_id=AWS_ACCESS_KEY_ID,
                                    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                                    region_name=AWS_REGION)

        try:
            response = textract_client.analyze_document(
                Document={'Bytes': image_file_object.read()},
                FeatureTypes=['FORMS', 'QUERIES'],
                QueriesConfig={'Queries': queries},
            )

            extracted_data = {}
            for block in response['Blocks']:
                if block['BlockType'] == 'QUERY':
                    alias = block['Query']['Alias']
                    query_id = block['Relationships'][0]['Ids'][0] if block.get('Relationships') else None
                    if query_id:
                        for result_block in response['Blocks']:
                            if result_block['BlockType'] == 'QUERY_RESULT' and result_block['Id'] == query_id:
                                extracted_data[alias] = result_block['Text']

            return extracted_data

        except ClientError as e:
            # Handle specific AWS exceptions
            error_code = e.response['Error']['Code']
            error_message = e.response['Error']['Message']
            print(f"Textract Client Error: {error_code} - {error_message}")
            # ... handle the error appropriately

        except Exception as e:
            # Handle general exceptions
            print(f"Error extracting text from image: {e}")
            # ... handle the error appropriately


    class Claim_pydantic(BaseModel):
        policy_number: str = Field(default_factory=str)
        claim_reported_by: str = Field(default_factory=str)
        loss_date_and_time: str = Field(default_factory=str, description="The loss date and time in YYYY/MM/DD HH:MM:SS "
                                                                        "format")
        loss_type: str = Field(default_factory=str, description="The cause for loss")
        street_number: int = Field(default_factory=int)
        street_name: str = Field(default_factory=str)
        loss_city: str = Field(default_factory=str)
        loss_state: str = Field(default_factory=str)
        loss_zip: int = Field(default_factory=int)
        loss_country: str = Field(default_factory=str)
        police_fire_contacted: bool = Field(default_factory=bool)
        report_number: str = Field(default_factory=str)
        loss_damage_description: str = Field(default_factory=str)
    claim_parser = JsonOutputParser(pydantic_object=Claim_pydantic)

    @classmethod
    def error_agent(extracted_text):
        prompt = PromptTemplate(
            template='''
            You are a expert data analyst who gets maximum information from provided content and also at the same time detects errors and
            spelling mistakes with in it.and also mismatches(like wrong state and coutry names or providing country name as state name) in the data and correct them!
            if the loss type is not specified go through the whole content you may find it somewhere in it, it can be any from this list [fire,storm,flood,earthquake,lightning,wind etc]!
            provide the output as specified.\n{format_instructions}\n{extracted_text} ''',
            input_variables=["extracted_text"],
            partial_variables={"format_instructions": Ai_utils.claim_parser.get_format_instructions()}
        )
        chain = prompt | Llama3_70b
        response = chain.invoke({
            "extracted_text": extracted_text
        })
        output_tokens, input_tokens, total_tokens = Ai_utils.returns_token_count(response)
        try:
            response = Ai_utils.claim_parser.invoke(response)
            return response, output_tokens, input_tokens, total_tokens
        except Exception as e:
            error_message = str(e)
            print("Error: ", error_message)
            # Improved Regex to capture more data types
            extracted_data = dict(re.findall(r'"([^"]+)":\s*("?[^",\n}]+"?|\d+)', error_message))
            print("extracted data in utils: ", extracted_data)
            if extracted_data:
                for key, value in extracted_data.items():
                    # Remove leading/trailing quotes
                    extracted_data[key] = value.strip('"')
                    # Convert to integer if possible
                    if extracted_data[key].isdigit():
                        extracted_data[key] = int(extracted_data[key])
                if extracted_data['street_number'] == 1 or extracted_data['street_number'] == 0:
                    extracted_data['street_number'] = "Null"
                return extracted_data, output_tokens, input_tokens, total_tokens
            else:
                return {"error": "Could not extract data from error message"}
            
    
    @classmethod
    def data_extraction_agent(cls, text):
        prompt = PromptTemplate(
            template='''
            You are a expert data analyst who retrieves import data from the given context, and also rectify any error or spelling mistakes in the provided data.The policy number usually have  10 or nearly 10 digits in it, so please search the whole text for it! I say never provide your own policy number, make sure to check the whole text atleast twice if you dont find the policy number!and also if you dont find a value for a key in the first time go through it again you may find it the second time! i think you know how street names will look like in general right!and provide the output as specified.\n{format_instructions}\n{text}.Make sure to retrieve the date and time of the loss in this format only: "YYYY/MM/DD HH:MM:SS" if it dont have any value even seconds provide 0 in place for example: "2024/05/27 16:46:00, please dont provide wrong format.Make sure to go through the whole text multiple times before providing the output as you are an expert analyst Your response could result in a big impact so be careful.** Main Point: The policy Number would be anything which is of size from 8-14 characters which consists ofcombination of letters, numbers etc.. So Please Carefully search for the Policy Number multiple times if you cant find it in the first time.
            ''',
            input_variables=["text"],
            partial_variables={"format_instructions": Ai_utils.claim_parser.get_format_instructions()}
        )
        chain = prompt | Llama3_1_8b
        response = chain.invoke({
            "text": text
        })
        output_tokens, input_tokens, total_tokens = Ai_utils.returns_token_count(response)
        try:

            response = Ai_utils.claim_parser.invoke(response)
            return response, output_tokens, input_tokens, total_tokens
        except Exception as e:
            error_message = str(e)
            extracted_data = Ai_utils.extract_data_from_error(error_message)
            if extracted_data:
                return extracted_data, output_tokens, input_tokens, total_tokens
            else:
                return f"got an error: {e}"
    
    @classmethod
    def get_accuracy(cls, input_text, response, details_to_extract):
        complete_text = "Fields to be extracted: \n"+ details_to_extract + "input_text: " + "\n" + str(input_text) + "\n\n" + "LLM Response: " + "\n" + str(response)

        class accuracy(BaseModel):
            accuracy: str = Field(default_factory=str)
            reason: str = Field(default_factory=str)

        accuracy_parser = JsonOutputParser(pydantic_object=accuracy)

        def accuracy_agent(text):
            prompt = PromptTemplate(
                template='''
                        You are an expert in evaluating the accuracy of LLM-generated outputs. You will be given:
                        1. Extraction criteria
                        2. Input text
                        3. The output generated by the LLM.

                        Your task is to calculate the accuracy of the output based on:
                        - Correctly extracted details
                        - Missing details
                        - Accuracy of extracted information.
                        Input:
                        \n{format_instructions}\n{extracted_text} 

                        Follow these rules to determine the accuracy:
                        ** If all fields are extracted correctly and match the input text, assign an accuracy between 98%-100%.
                        ** If all fields are extracted but the details differ entirely from the input (assumed or incorrect), assign an accuracy of 60-80% based on the error rate.
                        ** If fields are missing:
                        - 1 field missed: assign accuracy between 85-95%.
                        - 2-4 fields missed: assign accuracy between 75-85%.
                        - More than 4 fields missed: calculate accuracy proportionally based on the number of missing fields.

                        Output your evaluation strictly in the following JSON format:
                            "accuracy": <calculated_accuracy>,
                            "reason": "<brief explanation>"

                        Provide no additional text or commentary beyond the specified JSON format.
                        ** Make sure to give a correct accuracy value, don't simply give 95 everytime.
                    '''
                ,
                input_variables=["extracted_text"],
                partial_variables={"format_instructions": accuracy_parser.get_format_instructions()}
            )
            chain = prompt | Llama3_1_8b | accuracy_parser
            accuracy_response = chain.invoke({
                "extracted_text": text
            })
            return accuracy_response

        resp = accuracy_agent(complete_text)
        print("response: ", resp)
        resp_json_str = json.dumps(resp).replace('\\', '\\\\') 
        resp = json.loads(resp_json_str)
        return resp.get('accuracy', ''), resp.get('reason', '')
    
    @classmethod
    def process_pdf_image(cls, pdf_bytes, queries):
        """Processes a PDF by extracting the first page as an image and analyzing it."""
        if not pdf_bytes.getbuffer().nbytes:
            raise ValueError("The PDF file is empty or not readable.")

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")  # Open PDF from bytes
        page = doc[0]  # Select the first page (index 0)

        # Zoom if necessary (adjust zoom_x and zoom_y as needed)
        zoom_x = 3.0
        zoom_y = 3.0
        mat = fitz.Matrix(zoom_x, zoom_y)
        pix = page.get_pixmap(matrix=mat)

        # Convert pixmap to bytes
        image_bytes = pix.tobytes("png")

        # Analyze the image
        extracted_data = Ai_utils.extract_text_from_image(BytesIO(image_bytes), queries)

        return extracted_data  # Return the extracted data from the first page
    
    @classmethod
    def welcome_agent(cls, input):
        response = Llama3_70b_volatile.invoke(
            f"""
        You are an P&C (property and casualty) Insurance Expert Named "Ivan", who is currently working for the company : Innovon Technologies,
    Your work is solving queries of users related to p&c insurance, and also you will create a claim for them by taking information from them.
    But your task now is to provide a best and super attractive welcome message when our users come to use our AI Agent.
    Make the welcome message a little funny but formal, mention your name, your company name, and say a small and concise things about yourself.
    Dont make it more than 2 lines. Use emoji's and other attractive things in the response, when user says "hi".
    {input}
        """
        )
        return response.content

    @classmethod
    def process_extracted_data(cls, extracted_data: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        def convert_type(value: Any, target_type: type) -> Any:
            if value is None or value == "":
                return ""
            if target_type == int:
                try:
                    return int(float(value))
                except ValueError:
                    return ""
            if target_type == float:
                try:
                    return float(value)
                except ValueError:
                    return ""
            if target_type == str:
                return str(value)
            return value

        type_mapping = {
            "PolicyInfo": {
                "selectedPolicy": str,
                "policy_holder_FirstName": str,
                "policy_holder_LastName": str,
                "policy_holder_street_number": int,
                "policy_holder_street_name": str,
                "policy_holder_city": str,
                "policy_holder_state": str,
                "policy_holder_zip": int,
                "policy_holder_country": str,
                "policy_holder_mobile": str,
                "policy_holder_email": str,
                "policy_holder_occupation": str,
                "policy_holder_ssn": str,
                "policy_from_channel": str,
                "policy_associated_ic_id": str,
            },
            "PropertyInfo": {
                "residenceType": str,
                "constructionType": str,
                "otherconstructionType": str,
                "yearBuilt": int,
                "numberOfStories": int,
                "squareFootage": int,
                "heatingType": str,
                "plumbing_installed_year": int,
                "wiring_installed_year": int,
                "heating_installed_year": int,
                "roof_installed_year": int,
                "fireHydrantDistance": float,
                "fireStationDistance": float,
                "alternateHeating": str,
                "any_business_conducted_on_premises": str,
                "trampolineRamp": str,
                "subjectToFlood": str,
                "floodInsuranceRequested": str,
                "rentedToOthers": str,
                "CoverageLocation_street_number": int,
                "CoverageLocation_street_name": str,
                "CoverageLocation_city": str,
                "CoverageLocation_state": str,
                "CoverageLocation_zip": int,
                "CoverageLocation_country": str,
                "additionalInfo": str,
            },
            "AdditionalInfo": {
                "currentInsuranceCarrier": str,
                "currentPolicy": str,
                "effectiveDate": str,
                "current_policy_premium": float,
                "anyLossLast4Years": str,
                "mortgageeName": str,
                "mortgageeInstallmentAmount": float,
                "mortgageeStreetNumber": str,
                "mortgageeStreetName": str,
                "mortgageeCity": str,
                "mortgageeState": str,
                "mortgageeCountry": str,
                "mortgageeZip": str,
            },
            "Coverages": {
                "dwellingCoverage": float,
                "personalProperty": float,
                "personalLiabilityCoverage": float,
                "medicalPayments": float,
                "deductible": float,
            },
        }

        processed_data = {}

        for category, data in extracted_data.items():
            processed_category = {}
            for field in type_mapping[category]:
                value = data.get(field, "")
                converted_value = convert_type(value, type_mapping[category][field])
                processed_category[field] = converted_value
            processed_data[category] = processed_category

        return processed_data

    class InsuranceData(BaseModel):
        PolicyInfo: PolicyInfo
        PropertyInfo: PropertyInfo
        AdditionalInfo: AdditionalInfo
        Coverages: Coverages


    policy_parser = JsonOutputParser(pydantic_object=InsuranceData)
    
    @classmethod
    def policy_data_extraction_agent(cls, extracted_text):
        print("extracted_text: ", extracted_text)
        prompt = PromptTemplate(
            template='''Role: You are an Expert Data Analyst with over 20 years of experience in extracting and 
            interpreting critical information from unstructured text. Your task is to analyze the following extracted 
            text and provide a structured response according to the specified format:

            {format_instructions}

            Extracted Text:
            {extracted_text}

            Important Notes: 1. Pay close attention to the 'currentPolicy' field. It may contain the effective date 
            embedded within it. 2. When present, the policy number and effective date in the 'currentPolicy' field are 
            separated by an "&" symbol. 3. If you find the effective date in the 'currentPolicy' field, extract and 
            report it separately from the policy number, and only provide the policy Number in the 'currentPolicy' field
            we dont need effective date for that field.
            4. If you don't find any value for any field just provide empty string inplace (""). Dont provide 0.

            Please provide your analysis with high accuracy, ensuring all relevant information is correctly extracted and 
            categorized according to the given format instructions.''',

            input_variables=["extracted_text"],
            partial_variables={"format_instructions": Ai_utils.policy_parser.get_format_instructions()}
        )
        chain = prompt | Llama3_1_8b
        response = chain.invoke({
            "extracted_text": extracted_text
        })
        output_tokens, input_tokens, total_tokens = Ai_utils.returns_token_count(response)
        try:
            response = Ai_utils.policy_parser.invoke(response)
            return response, output_tokens, input_tokens, total_tokens
        except Exception as e:
            error_message = str(e)
            print('error Message: ', error_message)
            # Improved Regex to capture more data types
            extracted_data = dict(re.findall(r'"([^"]+)":\s*("?[^",\n}]+"?|\d+)', error_message))
            if extracted_data:
                for key, value in extracted_data.items():
                    # Remove leading/trailing quotes
                    extracted_data[key] = value.strip('"')
                    # Convert to integer if possible
                    if extracted_data[key].isdigit():
                        extracted_data[key] = int(extracted_data[key])
                return extracted_data, output_tokens, input_tokens, total_tokens
            else:
                return {"error": "Could not extract data from error message"}

    class medbill_pydantic(BaseModel):
        patient_info: patient_info
        guarantor_info: guarantor_info
        service_info: service_info

    medbill_parser = JsonOutputParser(pydantic_object=medbill_pydantic) 

    @classmethod
    def medbill_extraction_agent(cls, extracted_text):
        prompt = PromptTemplate(
            template='''
            Role: You are an Expert in Medical Data Extraction.
            Task: Extract the relevant details from the provided medical report.
            Instructions:
            Ensure that all monetary values are converted to float numbers and that the contact number is extracted as an integer.
            If any field is not present, use an empty string "" for text fields, 0.0 for float values, and 0 for integers.
            Do not make things up.
            Do not perform any calculations; simply extract the values as they appear in the text.
       
   
            Context: \n{extracted_text}  \n\n
            Follow these instructions to extract the information from the provided text: \n {format_instructions}
            ''',
            input_variables=["extracted_text"],
            partial_variables={"format_instructions": Ai_utils.medbill_parser.get_format_instructions()}
        )
        chain = prompt | Llama3_1_8b
        response = chain.invoke({
            "extracted_text": extracted_text
        })
        output_tokens, input_tokens, total_tokens = Ai_utils.returns_token_count(response)
        try:
            print("medBill pardser",Ai_utils.medbill_parser)
            response = Ai_utils.medbill_parser.invoke(response)

            return response, output_tokens, input_tokens, total_tokens
        except Exception as e:
            error_message = str(e)
            extracted_data = Ai_utils.extract_data_from_error(error_message)
            if extracted_data:
                return extracted_data, output_tokens, input_tokens, total_tokens
            else:
                return f"got an error: {e}"
            
 
    # Document types with their corresponding index keys
    DOCUMENT_TYPES_AND_KEYS = {
        "Insurance Application": [
            "policy_holder_name",
            "policy_holder_email",
            "policy_type",
            "property_address",
            "effective_date_of_coverage",
            "policy_holder_mobile"
        ],
        "Quotes and Proposals": [
            "Quote_or_Proposal_number",
            "Customer_name",
            "Effective_date",
            "Type_of_coverage",
            "Agent_name",
            "Underwriting_company_name"
        ],
        "Policy Declaration": [
            "Policy_number",
            "Policy_holder_name",
            "Policy_effective_date",
            "Policy_expiration_date",
            "Type_of_coverage",
            "Premium_amount"
        ],
        "Renewal Notice": [
            "Policy_number",
            "Policy_holder_name",
            "Policy_expiration_date",
            "Renewal_effective_date",
            "Policy_type"
        ],
        "Cancellation Notice": [
            "Policy_number",
            "Policy_holder_name",
            "Policy_effective_date",
            "Policy_expiration_date",
            "Policy_cancellation_date",
            "Reason_for_policy_cancellation"
        ],
        "First Notice of Loss (FNOL)": [
            "policy_number",
            "loss_date_and_time",
            "policy_effective_date",
            "policy_expiration_date",
            "loss_location"
        ],
        "Medical Bill": [
            "Patient_name",
            "Patient_ID",
            "Guarantor_name",
            "Guarantor_number",
            "Statement_date",
            "Service_date",          
            "Amount_due",            
            "Insurance_provider",    
            "Billing_date",       
            "Charge_amount" 
        ],
        "Credit Report": [
            "Policy_holder_name",
            "Policy_holder_mobile",
            "Policy_holder_email",
            "Insurance_score",
            "Credit_inquiry_type",
            "Credit_account_type"
        ],
        "Appraisal Report": [
            "Policy_holder_name",
            "Policy_number",
            "Claim_number",
            "Appraiser_name",
            "Appraisal_date",
            "Date_of_loss",
            "Appraisal_id_number"
        ],
        "Inspection Report": [
            "Policy_number",
            "Policy_holder_name",
            "Property_address",
            "Inspection_date",
            "Inspection_type",
            "Inspection_id_number"
        ]
    }

    DOCUMENT_TYPES = list(DOCUMENT_TYPES_AND_KEYS.keys())

    @classmethod
    def extract_headings(cls, text):
        """Extract potential headings from the document text."""
        try:
            if not isinstance(text, str):
                if isinstance(text, dict):
                    text = text.get('text', '') or text.get('content', '') or str(text)
                else:
                    text = str(text)

            lines = text.split('\n')
            headings = []
            for line in lines:
                line = line.strip()
                if line and (line.isupper() or line.endswith(':')) and len(line) < 100:
                    headings.append(line)
            return headings
        except Exception as e:
            print(f"Error extracting headings: {e}")
            return []

    @classmethod
    def calculate_key_match_score(cls, text, document_type):
        """Calculate how many index keys for a document type are found in the text."""
        if not text or not isinstance(text, str):
            return 0
        
        keys = cls.DOCUMENT_TYPES_AND_KEYS[document_type]
        text_lower = text.lower()
        matching_keys = sum(1 for key in keys if key.lower() in text_lower)
        return matching_keys / len(keys)

    @classmethod
    def documents_classification_agent(cls, extracted_file_text, file_name):
        print("inside agent")
        # document_classification_parser = PydanticOutputParser(pydantic_object=cls.DocumentClassification)        
        try:
            # Extract potential headings with error handling
            headings = cls.extract_headings(extracted_file_text)
            headings_str = "\n".join(f"- {heading}" for heading in headings) if headings else "No headers found"
            
            # Ensure extracted_file_text is a string
            if not isinstance(extracted_file_text, str):
                if isinstance(extracted_file_text, dict):
                    extracted_file_text = extracted_file_text.get('text', '') or extracted_file_text.get('content', '') or str(extracted_file_text)
                else:
                    extracted_file_text = str(extracted_file_text)

            # Calculate key match scores for each document type
            key_match_scores = {
                doc_type: cls.calculate_key_match_score(extracted_file_text, doc_type)
                for doc_type in cls.DOCUMENT_TYPES
            }

            # Find the document type with the highest key match score
            best_key_match = max(key_match_scores.items(), key=lambda x: x[1])
            
            prompt = PromptTemplate(
                template="""You are an expert document classifier specialized in insurance documents. 
                            File Name: {file_name}

                            Document Headings Found:
                            {headings}

                            Key Match Analysis:
                            {key_match_analysis}

                            Based on the file name, headings, key matches, and document content provided, please:

                            1. Analyze the following elements:
                            - File name patterns and keywords
                            - Document headings and structure
                            - Presence of index keys
                            - Full content details
                            
                            2. Classify it into one of the following categories:
                            {document_types}

                            3. Provide your classification in the following JSON format:
                            {{
                                "document_classification_name": "exact category name from the list",
                                "confidence_score": confidence score between 0 and 1
                            }}

                            Consider:
                            - The file name and any relevant headings as strong indicators
                            - The presence of index keys specific to each document type
                            - If a high percentage of index keys for a specific document type are found, that's a strong indicator

                            If the document doesn't clearly match any category, classify as "UnClassified Document" with an appropriate confidence score.

                            Document Content:
                            ---
                            {extracted_file_text}
                            ---

                            Classification (in JSON format):
                            """,
                            input_variables=["document_types", "extracted_file_text", "file_name", "headings", "key_match_analysis"]
                            )

            # Prepare the document types as a formatted string
            document_types_str = "\n".join(f"- {doc_type}" for doc_type in cls.DOCUMENT_TYPES)
            
            # Prepare key match analysis string
            key_match_analysis = "\n".join(
                f"- {doc_type}: {score:.2%} of index keys found"
                for doc_type, score in key_match_scores.items()
                if score > 0
            )
            if not key_match_analysis:
                key_match_analysis = "No significant index key matches found"

            # Invoke the LLM
            chain = prompt | Llama3_1_70b
            response = chain.invoke({
                "document_types": document_types_str,
                "extracted_file_text": extracted_file_text,
                "file_name": file_name,
                "headings": headings_str,
                "key_match_analysis": key_match_analysis
            })
            # print("LLM response:", response)            
            # Calculate tokens
            output_tokens, input_tokens, total_tokens = cls.returns_token_count(response)

            try:
                # Handle AIMessage response
                if hasattr(response, 'content'):
                    response_text = response.content
                else:
                    response_text = str(response)

                # Extract JSON from the response
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if not json_match:
                    # If no JSON found, check if the filename contains any document type
                    # or use the best key match if it's significant
                    if best_key_match[1] >= 0.5:  # If 50% or more keys match
                        return {
                            "document_classification_name": best_key_match[0],
                            "confidence_score": best_key_match[1]
                        }, output_tokens, input_tokens, total_tokens
                    
                    for doc_type in cls.DOCUMENT_TYPES:
                        if doc_type.lower() in file_name.lower():
                            return {
                                "document_classification_name": doc_type,
                                "confidence_score": 0.8  # High confidence due to filename match
                            }, output_tokens, input_tokens, total_tokens
                    raise ValueError("No JSON found in response")
                
                response_json = json.loads(json_match.group())

                # Validate the classification
                classification = response_json["document_classification_name"].strip()
                confidence = float(response_json.get("confidence_score", 0.0))

                # If classification is not in allowed types, check key matches and filename
                if classification not in cls.DOCUMENT_TYPES:
                    if best_key_match[1] >= 0.5:  # If 50% or more keys match
                        classification = best_key_match[0]
                        confidence = best_key_match[1]
                    else:
                        for doc_type in cls.DOCUMENT_TYPES:
                            if doc_type.lower() in file_name.lower():
                                classification = doc_type
                                confidence = 0.8  # High confidence due to filename match
                                break
                        else:  # If no match found in filename
                            classification = "UnClassified Document"
                            confidence = max(confidence, 0.3)

                return {
                    "document_classification_name": classification,
                    "confidence_score": confidence
                }, output_tokens, input_tokens, total_tokens

            except (json.JSONDecodeError, KeyError, ValueError) as e:
                print(f"Error parsing LLM response: {e}")
                # Check key matches before falling back to filename
                if best_key_match[1] >= 0.5:  # If 50% or more keys match
                    return {
                        "document_classification_name": best_key_match[0],
                        "confidence_score": best_key_match[1]
                    }, output_tokens, input_tokens, total_tokens
                
                # Check filename for document type
                for doc_type in cls.DOCUMENT_TYPES:
                    if doc_type.lower() in file_name.lower():
                        return {
                            "document_classification_name": doc_type,
                            "confidence_score": 0.8  # High confidence due to filename match
                        }, output_tokens, input_tokens, total_tokens
                
                return {
                    "document_classification_name": "UnClassified Document",
                    "confidence_score": 0.3
                }, output_tokens, input_tokens, total_tokens

        except Exception as e:
            print(f"Classification error: {str(e)}")
            # Even in case of error, check key matches before filename
            if best_key_match[1] >= 0.5:  # If 50% or more keys match
                return {
                    "document_classification_name": best_key_match[0],
                    "confidence_score": best_key_match[1]
                }, 0, 0, 0
            
            # Check filename for document type
            for doc_type in cls.DOCUMENT_TYPES:
                if doc_type.lower() in file_name.lower():
                    return {
                        "document_classification_name": doc_type,
                        "confidence_score": 0.8  # High confidence due to filename match
                    }, 0, 0, 0
            
            return {
                "document_classification_name": "UnClassified Document",
                "confidence_score": 0.0
            }, 0, 0, 0

 
    MODEL_MAPPING = {
        "Insurance Application": InsuranceApplicationData,
        "Quotes and Proposals": QuotesAndProposalsData,
        "Policy Declaration": PolicyDeclarationData,
        "Renewal Notice": RenewalNoticeData,
        "Cancellation Notice": CancellationNoticeData,
        "First Notice of Loss (FNOL)": FirstNoticeOfLossData,
        "Medical Bill": MedicalBillData,
        "Credit Report": CreditReportData,
        "Appraisal Report": AppraisalReportData,
        "Inspection Report": InspectionReportData,
    }
    # Document-specific prompt templates with detailed instructions
    PROMPT_TEMPLATES = {
        "Insurance Application": """You are an expert insurance document analyzer. Your task is to extract specific information from an insurance application.
                Context: This is an insurance application document that contains details about a potential policy holder.
                Required Information to Extract:
                1. Policy Holder's Full Name (including any middle names or suffixes)
                2. Email Address (in standard email format)
                3. Type of Insurance Policy (e.g., home, auto, life, etc.)
                4. Property Address (complete address including city, state, zip)
                5. Effective Date of Coverage (in YYYY-MM-DD format)
                6. Mobile Number (in standard format)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - For dates, standardize to YYYY-MM-DD format
                - If a field is not found, use "Not Found" as the value
                - Maintain proper formatting for email and phone numbers
                - Calculate confidence score based on completeness and clarity of extracted data

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Double-check all extracted values for accuracy
                - Ensure all required fields are included
                - Format dates and numbers consistently
                - Include confidence scores for uncertain extractions""",

        "Quotes and Proposals": """You are an expert insurance document analyzer. Your task is to extract specific information from an insurance quote or proposal.
                Context: This document contains details about an insurance quote or proposal provided to a potential customer.
                Required Information to Extract:
                1. Quote/Proposal Number (exact identifier)
                2. Customer Name (full name)
                3. Effective Date (in YYYY-MM-DD format)
                4. Type of Coverage (specific insurance type)
                5. Agent Name (full name of insurance agent)
                6. Underwriting Company Name (complete legal name)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Include complete company names
                - Calculate confidence score based on data clarity

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify all extracted information
                - Check for alternative field names/locations
                - Include any relevant qualifiers
                - Note any ambiguities in the extraction""",

        "Policy Declaration": """You are an expert insurance document analyzer. Your task is to extract specific information from a Policy Declaration document.
                Context: This document officially outlines the terms and conditions of an insurance policy. Accuracy is crucial as this is a legal document.
                Required Information to Extract:                
                1. Policy Number (exact identifier)
                2. Policy Holder's Full Name (including any middle names or suffixes)
                3. Policy Effective Date (in YYYY-MM-DD format)
                4. Policy Expiration Date (in YYYY-MM-DD format)
                5. Type of Coverage (specific insurance type)
                6. Premium Amount (exact amount with currency $)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Maintain proper formatting for currency amounts
                - Calculate confidence score based on completeness and clarity of extracted data

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Double-check all extracted values for accuracy
                - Verify policy terms and conditions
                - Format currency and numbers consistently
                - Include confidence scores for uncertain extractions""",

        "Renewal Notice": """You are an expert insurance document analyzer. Your task is to extract specific information from a Renewal Notice.
                Context: This document informs the policyholder about the renewal of their insurance policy and any changes to terms or premiums.
                Required Information to Extract:
                1. Policy Number (exact identifier)
                2. Policy Holder's Full Name (including any middle names or suffixes)
                3. Policy Expiration Date (in YYYY-MM-DD format)
                4. Renewal Effective Date (in YYYY-MM-DD format)
                5. Policy Type (specific insurance type)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Highlight any premium or coverage changes
                - Calculate confidence score based on data clarity

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Compare old and new premium amounts
                - Note any changes in coverage terms
                - Verify payment deadlines
                - Include confidence scores for uncertain extractions""",

        "Cancellation Notice": """You are an expert insurance document analyzer. Your task is to extract specific information from a Cancellation Notice.
                Context: This document informs the policyholder about the cancellation of their insurance policy. This is a critical document requiring high accuracy.
                Required Information to Extract:
                1. Policy Number (exact identifier)
                2. Policy Holder's Full Name (including any middle names or suffixes)
                3. Policy Effective Date (in YYYY-MM-DD format)
                4. Policy Expiration Date (in YYYY-MM-DD format)
                5. Policy Cancellation Date (in YYYY-MM-DD format)
                6. Reason for Policy Cancellation (specific details)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Pay special attention to cancellation reasons
                - Calculate confidence score based on completeness and clarity

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify all cancellation details carefully
                - Check for any grace period information
                - Note any reinstatement conditions
                - Include confidence scores for uncertain extractions""",

        "First Notice of Loss (FNOL)": """You are an expert insurance document analyzer. Your task is to extract specific information from a First Notice of Loss (FNOL).
                Context: This document reports an initial claim for a loss covered by an insurance policy. Accuracy and detail are crucial for claims processing.
                Required Information to Extract:
                1. Policy Number (exact identifier)
                2. Loss Date and Time (YYYY-MM-DD HH:MM or best available format)
                3. Policy Effective Date (in YYYY-MM-DD format)
                4. Policy Expiration Date (in YYYY-MM-DD format)
                5. Loss Location (complete address including city, state, zip)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize dates to YYYY-MM-DD HH:MM format
                - Use "Not Found" for missing fields
                - Maintain proper formatting for contact information
                - Calculate confidence score based on completeness and detail

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify incident details thoroughly
                - Check for emergency response information
                - Note any witness information
                - Include confidence scores for uncertain extractions""",

        "Medical Bill": """You are an expert insurance document analyzer. Your task is to extract specific information from a Medical Bill.
                Context: This document details charges for medical services provided to a patient. Accuracy is essential for proper claims processing.
                Required Information to Extract:
                1. Patient's Full Name (including any middle names or suffixes)
                2. Patient ID/Medical Record Number (exact identifier)
                3. Guarantor's Full Name (including any middle names or suffixes)
                4. Guarantor Number/Account Number (exact identifier)
                5. Statement Date (in YYYY-MM-DD format)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Maintain proper formatting for currency amounts
                - Calculate confidence score based on completeness and clarity

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify all Patient and Guarantor Names
                - Check for multiple service dates
                - Check for multiple Account, Record Numbers
                - Confirm provider information
                - Include confidence scores for uncertain extractions""",

        "Credit Report": """You are an expert insurance document analyzer. Your task is to extract specific information from a Credit Report used for insurance purposes.
                Context: This document provides credit information used for insurance underwriting and risk assessment. The information is highly sensitive and requires careful handling and accurate extraction.
                Required Information to Extract: 
                1. Policy Holder's Full Name (including any middle names or suffixes)
                2. Policy Holder Mobile (in standard email format)
                3. Policy Holder Email (valid email format)
                4. Insurance Score (numeric value exact indetifier)
                5. Credit Inquiry Type (soft/hard etc.. inquiry exact indetifier)
                6. Credit Account Type (revolving/installment/mortgage etc.. type exact indetifier )


                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Use "Not Found" for missing fields
                - Maintain proper formatting for contact information
                - Ensure numeric scores are extracted precisely
                - Calculate confidence score based on completeness and clarity
                - Handle sensitive information with appropriate care

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify the authenticity of the credit report source
                - Double-check all numeric scores and dates
                - Confirm proper formatting of contact information
                - Note any alerts or fraud indicators
                - Pay special attention to inquiry types and account types
                - Include confidence scores for uncertain extractions
                - Include confidence scores for uncertain extractions""",

        "Appraisal Report": """You are an expert insurance document analyzer. Your task is to extract specific information from an Appraisal Report.
                Context: This document provides a professional valuation of damages or losses. Accuracy is crucial for claims settlement.
                Required Information to Extract:
                1. Policy Holder's Full Name (including any middle names or suffixes)
                2. Policy Number (exact identifier)
                3. Claim Number (exact identifier)
                4. Appraiser's Full Name (including any middle names or suffixes)
                5. Appraisal Date (in YYYY-MM-DD format)
                6. Date of Loss (in YYYY-MM-DD format)
                7. Appraisal ID Number (exact identifier)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Maintain proper formatting for currency amounts
                - Calculate confidence score based on detail and completeness

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify multiple ID, Acoount Numbers for better identification
                - Check for all the Names and Dates
                - Note any special conditions or exceptions
                - Include confidence scores for uncertain extractions""",

        "Inspection Report": """You are an expert insurance document analyzer. Your task is to extract specific information from an Inspection Report.
                Context: This document details the findings of a property inspection for insurance purposes. Thoroughness is essential.
                Required Information to Extract:
                1. Policy Number (exact identifier)
                2. Policy Holder's Full Name (including any middle names or suffixes)
                3. Property Address (complete address including city, state, zip)
                4. Inspection Date (in YYYY-MM-DD format)
                5. Inspection Type (purpose/scope)
                6. Inspection ID Number (exact identifier)

                Document Content:
                ---
                {extracted_text}
                ---

                Important Instructions:
                - Extract exact values as they appear in the document
                - Standardize all dates to YYYY-MM-DD format
                - Use "Not Found" for missing fields
                - Maintain proper formatting for measurements
                - Calculate confidence score based on completeness and detail

                Please provide the extracted information in the following JSON format:
                {format_instructions}

                Remember to:
                - Verify all the Identification Numbers like ID etc..
                - Be causious about Inspection and Policy dates.
                - Check for addresses including the whole address fields
                - Check for the Multiple dates for better identification
                - Include confidence scores for uncertain extractions""",
    }


    @classmethod
    def indexdata_extraction_agent(cls, classification_type: str, extracted_text: str):
        """
        Main method for extracting indexed data from documents
        
        Args:
            classification_type: Type of document being processed
            extracted_text: Text content of the document
            
        Returns:
            tuple: (extracted_data, output_tokens, input_tokens, total_tokens)
        """
        try:
            # print(f"Processing document type: {classification_type}")           
            model = cls.MODEL_MAPPING.get(classification_type)
            if not model:
                raise ValueError(f"Unsupported document type: {classification_type}")
            # print(f"Using model: {model.__name__}", model)
            parser = JsonOutputParser(pydantic_object=model)
            prompt_template = cls.PROMPT_TEMPLATES.get(
                classification_type, 
            )
            # print(" in prompt template", prompt_template)
            if not prompt_template:
                raise ValueError(f"No prompt template available for {classification_type}")

            prompt = PromptTemplate(
                template=prompt_template,
                input_variables=["extracted_text"],
                partial_variables={"format_instructions": parser.get_format_instructions()}
            )
            # print(" in prompt", prompt)
            chain = prompt | Llama3_1_8b
            response = chain.invoke({
            "extracted_text": extracted_text
            })
            # print(" LLM response after extraction", response)
            output_tokens, input_tokens, total_tokens = Ai_utils.returns_token_count(response)
            try:
                formatedResponse = parser.invoke(response)
                # print("formatedResponse",formatedResponse)
                return formatedResponse, output_tokens, input_tokens, total_tokens

            except Exception as e:
                print(f"Error parsing extraction response: {e}")
                # Return empty data with appropriate structure
                empty_data = {
                    "extracted_date": datetime.now().strftime("%Y-%m-%d"),
                    "confidence_score": 0.0
                }
                empty_data.update({field: "Not Found" for field in model.__fields__ if field not in empty_data})
                return empty_data, output_tokens, input_tokens, total_tokens

        except Exception as e:
            print(f"Error in index data extraction: {e}")
            empty_data = {
                "extracted_date": datetime.now().strftime("%Y-%m-%d"),
                "confidence_score": 0.0
            }
            if classification_type in cls.MODEL_MAPPING:
                empty_data.update({
                    field: "Not Found" 
                    for field in cls.MODEL_MAPPING[classification_type].__fields__ 
                    if field not in empty_data
                })
            return empty_data, 0, 0, 0
    
 
    

class Policy_utils:

    @classmethod
    def generate_next_quote_number(cls, latest_quote_number):
        current_year = str(datetime.today().strftime("%Y"))[2:]
        if latest_quote_number:
            prefix = latest_quote_number[:2]  # Extract prefix (e.g., "HO")
            numeric_part = int(latest_quote_number[4:])  # Extract numeric part
        else:
            prefix = "HO"  # Default prefix
            numeric_part = 1000  # Starting number
        next_numeric_part = numeric_part + 1 if numeric_part < 999999 else 1
        next_quote_number = f"{prefix}{current_year}{str(next_numeric_part).zfill(6)}"
        return next_quote_number
    
    @classmethod
    def generate_random_quote_amount(cls):
        """Generates a random number between 500 and 2000, rounded to the nearest 10."""

        number = random.randint(50, 200) * 10
        return number
    
    @classmethod
    def add_policy_info_to_db(cls, policy_info, property_info, coverage_and_additional_info):
        client, db = MongoDB.get_mongo_client_Policy_intake()
        policy_information_collection = db['policy_holder_information']
        property_information_collection = db['property_information']
        coverage_and_additional_information_collection = db['coverage_and_additional_information']
        policy_number = policy_info.get('policy_number', '')
        policy_holder_name = policy_info.get('claimant_FirstName', '') + ' ' + policy_info.get('claimant_LastName', '')
        policy_type = policy_info.get('selectedPolicy', '')
        email = policy_info.get('claimant_email', '')
        try:
            encrypted_policy_info = Authentication.encrypt_data(policy_info)
            encrypted_property_info = Authentication.encrypt_data(property_info)
            encrypted_coverage_and_additional_info = Authentication.encrypt_data(coverage_and_additional_info)
            policy_information_collection.insert_one(encrypted_policy_info)
            property_information_collection.insert_one(encrypted_property_info)
            coverage_and_additional_information_collection.insert_one(encrypted_coverage_and_additional_info)
            Emails.send_policy_intake_success_mail(policy_number, policy_holder_name, policy_type, email)
            return "Policy created successfully"
        except Exception as e:
            return f"Got an unexpected error when adding policy to Database: {str(e)}"


class Administration_utils:
    @classmethod
    def delete_draft_data_from_DB(cls, user_email, portal_type):
            """
            Deletes draft data from the 'Portals_Draft' collection based on user email and portal type.

            Args:
                user_email (str): The email of the user whose draft data should be deleted.
                portal_type (str): The type of portal (e.g., "claim", "quote").

            Returns:
                dict: A dictionary containing the result of the deletion operation.
                    The dictionary will have the following keys:
                    - 'acknowledged': True if the operation was acknowledged by MongoDB, False otherwise.
                    - 'deleted_count': The number of documents deleted.
            """

            try:
                client, db = MongoDB.get_mongo_client_Administration()
                draft_data_collection = db['Portals_Draft']
                existing_draft = draft_data_collection.find_one({
                    'user_email': user_email,
                    'portal_type': portal_type
                })
                if not existing_draft:
                    return {
                        'ok': 1.0, 
                        'n': 0,    
                        'nModified': 0,
                        'message': 'No draft found'
                    }
                delete_result = draft_data_collection.delete_many({
                    'user_email': user_email,
                    'portal_type': portal_type
                })
                return delete_result.raw_result 
            except Exception as e:
                print(f"Error deleting draft data: {e}")
                return { 
                    'ok': 0.0, 
                    'n': 0,     
                    'error': str(e), 
                }